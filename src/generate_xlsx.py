from pathlib import Path
import random
import re

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.config import DIST_DIR
from src.utils.get_docx_data import docx_data_to_ls

"""
- 绿色字体 用于强调：请勿删除这个公式单元格。且它的下方均应通过'下拉填充'产生，而非手填
- Random/123 仅作为 用户名/id数字 示例，可实际调整。
"""
GREEN_FONT = Font(color="00C000")   # 绿色强调：公式单元格勿删
BOLD_FONT = Font(bold=True)

# 使用说明文本（保留原有内容）
INSTRUCTIONS = [
    "1. 绿色字体单元格为公式单元格，请勿删除！其下方内容需通过下拉填充生成（勿手填）",
    "2. 有新参赛者时，请务必先在第一个工作表填写 '用户名/id数字' 和 选题情况，再到后面的工作表去批改具体题目。选题情况中：'1' 表示选该题，空表示未选",
    "3. 批改完成情况中：'✔' 表示已完成批改，'✘' 表示未完成。自行下拉即可查看批改情况。",
    "4. 批改具体题目时：满分值行用于填写各题满分标准（按自设分值修改）；用户分行对应实际得分。"
]


def _clean_sheet_name(raw_title: str) -> str:
    """去除结尾的'（XX分）'并替换在线表格不允许的字符。"""
    # 去掉“（25分）”等分值后缀
    clean = re.sub(r"（\d+分）$", "", raw_title).strip()
    # 替换非法字符
    clean = re.sub(r"[:：\\/／?？*＊\[\]［］]", "_", clean)
    return clean


def _build_info_sheet(
    wb: Workbook,
    sheet_names: list[str],
    sheet_mapping: list[tuple[str, str]],
    question_data: dict[str, list[tuple[str, int]]]
) -> tuple[Worksheet, int, int]:
    """
    构建“答卷基本信息”工作表，包含用户ID、选题情况、批改完成情况和使用说明。
    返回：工作表对象、选题区域结束列、批改区域结束列。
    """
    ws = wb.create_sheet("答卷基本信息")

    # 基础信息行
    ws["A2"] = "用户名/id"
    ws["A2"].font = BOLD_FONT
    ws["A3"] = "满分值"
    ws["A4"] = "Random/123"   # 示例用户名，可替换

    num_sheets = len(sheet_names)
    start_sel = 2  # 选题情况从 B 列开始
    end_sel = start_sel + num_sheets - 1

    # 选题情况标题
    ws.merge_cells(start_row=1, start_column=start_sel, end_row=1, end_column=end_sel)
    sel_cell = ws.cell(row=1, column=start_sel, value="选题情况")
    sel_cell.font = BOLD_FONT
    sel_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 批改完成情况标题（紧接其后）
    start_corr = end_sel + 1
    end_corr = start_corr + num_sheets - 1
    ws.merge_cells(start_row=1, start_column=start_corr, end_row=1, end_column=end_corr)
    corr_cell = ws.cell(row=1, column=start_corr, value="批改完成情况")
    corr_cell.font = BOLD_FONT
    corr_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 第2行填入清洗后的题名（左对齐）
    for j, name in enumerate(sheet_names):
        ws.cell(row=2, column=start_sel + j, value=name).alignment = Alignment(horizontal="left")
        ws.cell(row=2, column=start_corr + j, value=name).alignment = Alignment(horizontal="left")

    # 示例选题数据：满分行填1，用户行随机1或空。填入1表示选该题，不填任何东西（即空单元格）表示不选题
    for j in range(num_sheets):
        ws.cell(row=3, column=start_sel + j, value=1)
        ws.cell(row=4, column=start_sel + j, value=random.choice([1, ""]))

    # 批改完成情况公式
    for j, name in enumerate(sheet_names):
        col_letter = get_column_letter(start_sel + j)
        selection_ref = f"{col_letter}3"
        sheet_b2_ref = f"'{name}'!B2"
        formula = f'=IF(AND({selection_ref}<>"", ISBLANK({sheet_b2_ref})), "✘", "✔")'
        ws.cell(row=3, column=start_corr + j, value=formula).font = GREEN_FONT

    # 使用说明区域（跨多列）
    inst_start = end_corr + 2
    inst_end = inst_start + 12
    ws.merge_cells(start_row=1, start_column=inst_start, end_row=1, end_column=inst_end)
    inst_title = ws.cell(row=1, column=inst_start, value="使用说明")
    inst_title.font = BOLD_FONT
    inst_title.alignment = Alignment(horizontal="center", vertical="center")

    for i, text in enumerate(INSTRUCTIONS, start=2):
        ws.merge_cells(start_row=i, start_column=inst_start, end_row=i, end_column=inst_end)
        cell = ws.cell(row=i, column=inst_start, value=text)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    return ws, end_sel, end_corr


def _build_question_sheets(
    wb: Workbook,
    sheet_mapping: list[tuple[str, str]],
    question_data: dict[str, list[tuple[str, int]]]
) -> list[str]:
    """为每道大题创建独立工作表，填入小题编号和满分。返回清洗后的表名列表。"""
    sheet_names: list[str] = []
    for raw_title, clean_title in sheet_mapping:
        ws = wb.create_sheet(clean_title)
        sheet_names.append(clean_title)

        # 基础公式区域
        ws["A1"] = "用户名/id"
        ws["A1"].font = BOLD_FONT
        ws["A2"] = "满分值"
        ws["A3"] = "=答卷基本信息!A4"  # 引用用户ID
        ws["A3"].font = GREEN_FONT

        ws["B1"] = "总分＼小题号"
        ws["B2"] = "=SUM(C2:IV2)"  # 总分公式
        ws["B2"].font = GREEN_FONT

        # 填入小题信息（编号与满分）
        questions = question_data.get(raw_title, [])
        for idx, (q_num, q_score) in enumerate(questions):
            col = 3 + idx  # 从 C 列开始
            ws.cell(row=1, column=col, value=q_num)
            ws.cell(row=2, column=col, value=q_score)

    return sheet_names


def _build_summary_sheet(wb: Workbook, sheet_names: list[str]) -> None:
    """创建“总分统计”表：全卷总分、最高三题总分、排名公式。"""
    ws = wb.create_sheet("总分统计")

    ws["A1"] = "用户名/id"
    ws["A1"].font = BOLD_FONT
    ws["A2"] = "满分值"
    ws["A3"] = "=答卷基本信息!A4"
    ws["A3"].font = GREEN_FONT

    # 全卷总分（满分与实际得分）
    ws["B1"] = "全卷总分"
    ws["B2"] = "=" + "+".join([f"'{s}'!B2" for s in sheet_names])
    ws["B2"].font = GREEN_FONT
    ws["B3"] = "=" + "+".join([f"'{s}'!B3" for s in sheet_names])
    ws["B3"].font = GREEN_FONT

    # 最高三题总分（只计算得分 >0 的选题）
    ws["C1"] = "最高三题总分"
    n = len(sheet_names)
    choose_index = "{" + ",".join(str(i) for i in range(1, n + 1)) + "}"
    refs_b2 = ",".join(f"'{s}'!B2" for s in sheet_names)
    refs_b3 = ",".join(f"'{s}'!B3" for s in sheet_names)
    cond_b2 = f"IF(CHOOSE({choose_index},{refs_b2})>0,CHOOSE({choose_index},{refs_b2}))"
    cond_b3 = f"IF(CHOOSE({choose_index},{refs_b3})>0,CHOOSE({choose_index},{refs_b3}))"

    ws["C2"] = "=" + " + ".join([
        f"IFERROR(LARGE({cond_b2},1),0)",
        f"IFERROR(LARGE({cond_b2},2),0)",
        f"IFERROR(LARGE({cond_b2},3),0)"
    ])
    ws["C3"] = "=" + " + ".join([
        f"IFERROR(LARGE({cond_b3},1),0)",
        f"IFERROR(LARGE({cond_b3},2),0)",
        f"IFERROR(LARGE({cond_b3},3),0)"
    ])
    ws["C2"].font = GREEN_FONT
    ws["C3"].font = GREEN_FONT

    # 排名公式：先比最高三题总分，再比全卷总分
    ws["D1"] = "排名"
    ws["D3"] = (
        "=SUMPRODUCT(($C$3:$C$1000>C3)*1)+"
        "SUMPRODUCT(($C$3:$C$1000=C3)*($B$3:$B$1000>B3)*1)+1"
    )
    ws["D3"].font = GREEN_FONT


def _apply_center_alignment(
    wb: Workbook,
    inst_start: int, inst_end: int,
    sel_start: int, sel_end: int,
    corr_start: int, corr_end: int
) -> None:
    """统一设置居中对齐，但保留使用说明和题名行的左对齐。"""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if ws.title == "答卷基本信息":
                    # 使用说明区域保持左对齐
                    if inst_start <= cell.column <= inst_end:
                        continue
                    # 第2行的题名需左对齐
                    if cell.row == 2 and (
                        sel_start <= cell.column <= sel_end or
                        corr_start <= cell.column <= corr_end
                    ):
                        continue
                cell.alignment = Alignment(horizontal="center", vertical="center")


def build_workbook(question_data: dict[str, list[tuple[str, int]]]) -> Workbook:
    """主构建函数：组装完整的工作簿（不含保存）。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 移除默认空白工作表

    # 工作表名清洗映射
    sheet_mapping: list[tuple[str, str]] = [
        (raw, _clean_sheet_name(raw)) for raw in question_data
    ]
    clean_names = [clean for _, clean in sheet_mapping]

    # 依次构建各模块
    _, sel_end, corr_end = _build_info_sheet(wb, clean_names, sheet_mapping, question_data)
    _build_question_sheets(wb, sheet_mapping, question_data)
    _build_summary_sheet(wb, clean_names)

    # 对齐设置
    num = len(clean_names)
    sel_start = 2
    sel_final = sel_start + num - 1
    corr_start = sel_final + 1
    corr_final = corr_start + num - 1
    inst_start = corr_final + 2
    inst_end = inst_start + 12
    _apply_center_alignment(wb, inst_start, inst_end, sel_start, sel_final, corr_start, corr_final)

    return wb


def main() -> None:
    """
    程序入口：从 docx 读取题目信息 -> 生成 xlsx -> 保存到 dist 目录。
    """
    question_data = docx_data_to_ls()
    wb = build_workbook(question_data)
    DIST_DIR.mkdir(parents=True, exist_ok=True)
    output_path = DIST_DIR / "final.xlsx"
    wb.save(output_path)
    print(f"✓ 文件已保存到 {output_path}")


if __name__ == "__main__":
    main()
